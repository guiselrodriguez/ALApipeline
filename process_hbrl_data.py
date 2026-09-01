#!/usr/bin/env python3
"""
HBRL house data pipeline.

Four subcommands:

  sensors           Combines a house's raw instrument files (Atmocube,
                     AirAssure, Geocene, Hobo, Anemometer, Kestrel, Aranet)
                     into ONE table per house:
                       house_id, phase, date, timestamp_utc, instrument,
                       variable, value, unit, qc_flag, sensor_id
                     All timestamps are stored in TRUE UTC. 'phase' (pre/post)
                     comes from House_charac's Visit 1/2/3 dates -- pre ends
                     30 min before Visit 2, post starts 1 hour after Visit 2
                     (the visit itself is excluded as a disruption buffer).

  house-charac      Builds a static per-house metadata table from the
                     House_charac spreadsheet (one row per house).

  ogawa             Builds the Ogawa NO2 table from the Ogawa_raw tab
                     (house, phase, location, value -- one row per reading).

  verify-timezones  Prints raw-vs-converted-UTC timestamps per instrument,
                     for manual review, using the SAME parser functions the
                     sensors command uses (not a re-implementation).

Usage:
    python3 process_hbrl_data.py sensors <house_charac.xlsx> <pre_folder_or_zip> <post_folder_or_zip> [--sensor-ids <file>] [--output <folder>]
    python3 process_hbrl_data.py house-charac <house_charac.xlsx> [output_folder] [--house H1 H2 ...]
    python3 process_hbrl_data.py ogawa <values.xlsx> <placement.xlsx> [output_folder] [--house H1 ...]
    python3 process_hbrl_data.py verify-timezones <folder_or_zip> [--samples 3] [--output <file>]

File naming: files can come in 3 different naming styles, and the script
checks for all of them so old and new files both work in the same run:
  1. NEW:    h1_pre_atmocube.csv / h1_post_airassure_pt1.csv
  2. OLD:    h1_w1_atmocube.csv / h1_w2_kestrel.csv
  3. LEGACY: house1_week2_geocene_left.json
Whatever matches is just used for labeling, it doesn't decide
pre/post on its own. That's always figured out per row using House_charac's
Visit dates, no matter what the filename says. Inputs can be zip files or
already-extracted folders, either works.

Some folders get skipped entirely no matter what's in them: anything with
"processed", "metrics", or "mission_logs" in the path. These have turned out
to hold someone's own already-processed output (can have the exact same
filename as the real raw file) or raw device export dumps whose files can
coincidentally match our naming pattern.

Timezone handling:
 - Atmocube, Hobo, and AirAssure all label their own timezone in the raw
   file, so those just get converted straight to UTC.
 - Geocene's raw timestamps normally end in 'Z' (UTC) -- when that's there,
   it converts directly. If not, it's treated as local Oregon time (PDT)
   and converted. This is checked per file rather than assumed.
 - Kestrel, Aranet, and Anemometer are local Oregon time (PDT, UTC-7), so
   that's the offset used to convert them to UTC.
 - House_charac's Visit dates use the same PDT conversion.

Some AirAssure/Geocene files are missing entirely for a house/period and
only exist as the device's own raw export zip, not the clean file this
pipeline normally reads. For Geocene there's a fallback that reads that
zip directly when the clean file isn't found.
"""

import argparse
import csv
import glob
import gzip
import json
import os
import re
import shutil
import sys
import tempfile
import zipfile
from collections import defaultdict
from datetime import datetime, timedelta, timezone

try:
    import openpyxl
except ImportError:
    openpyxl = None  # only needed for house-charac, ogawa, and --sensor-ids

PDT = timezone(timedelta(hours=-7))  # Local Oregon time


def fmt_time(dt):
    return dt.strftime("%I:%M %p").lstrip("0")


# Filename matching -- three conventions, tried in order


FNAME_RE_NEW = re.compile(r'^h(\d+)_(pre|post)_(.+)\.([^.]+)$', re.IGNORECASE)
FNAME_RE_OLD = re.compile(r'^h(\d+)_w(\d+)_(.+)\.([^.]+)$', re.IGNORECASE)
FNAME_RE_LEGACY = re.compile(r'^house(\d+)_week(\d+)_(.+)\.([^.]+)$', re.IGNORECASE)

# skip anything in these folders -- someone's own processed output, or raw
# device export dumps whose files can coincidentally match our naming pattern
EXCLUDED_PATH_KEYWORDS = ["processed", "metrics"]


def match_filename(base):
    """Returns (house_num, group_label, rest, ext) or None."""
    m = FNAME_RE_NEW.match(base)
    if m:
        house_num, phase, rest, ext = m.groups()
        return house_num, phase.lower(), rest, ext
    m = FNAME_RE_OLD.match(base)
    if m:
        house_num, week_num, rest, ext = m.groups()
        return house_num, f"w{week_num}", rest, ext
    m = FNAME_RE_LEGACY.match(base)
    if m:
        house_num, week_num, rest, ext = m.groups()
        return house_num, f"w{week_num}", rest, ext
    return None


# Checks a file is really raw AirAssure data and not an old output file that got mixed in by accident (happened once)
def _is_genuine_raw_airassure(path):
    try:
        with open(path, newline="") as f:
            for raw in f:
                line = raw.rstrip("\n")
                if not line.strip() or line.startswith("#"):
                    continue
                return line.startswith("Timestamp,System Status")
    except (FileNotFoundError, UnicodeDecodeError):
        return False
    return False


def _airassure_date_range(path):
    """Peeks first/last raw timestamp in an AirAssure file -- used to tell
    if a 'combined' file overlaps with the individual parts (duplicate,
    drop it) or covers a period the parts don't have (keep it)."""
    header = None
    first_dt, last_dt = None, None
    with open(path, newline="") as f:
        for raw in f:
            line = raw.rstrip("\n")
            if not line.strip() or line.startswith("#"):
                continue
            if line.startswith("Timestamp,System Status"):
                header = next(csv.reader([line]))
                continue
            if line.startswith("UTC,") or header is None:
                continue
            rec = dict(zip(header, next(csv.reader([line]))))
            raw_ts = rec.get("Timestamp")
            if not raw_ts:
                continue
            dt = _parse_airassure_timestamp(raw_ts)
            if dt is None:
                continue
            if first_dt is None:
                first_dt = dt
            last_dt = dt
    return first_dt, last_dt


def _ranges_overlap(a, b):
    return a[0] is not None and b[0] is not None and a[0] <= b[1] and b[0] <= a[1]


def discover_groups(folder):
    groups = defaultdict(dict)
    airassure_parts_by_house = defaultdict(list)
    airassure_combined_by_house = defaultdict(list)
    geocene_raw_sources = defaultdict(list)  # raw exports (zip OR already-extracted folder), tracked as a fallback

    for path in sorted(glob.glob(os.path.join(folder, "**", "*"), recursive=True)):
        if os.path.isdir(path):
            continue
        path_lower = path.lower()
        if any(kw in path_lower for kw in EXCLUDED_PATH_KEYWORDS):
            continue

        base = os.path.basename(path)

        # A raw Geocene export shows up either as a .zip, or already
        # unzipped with missions.csv sitting loose in the folder -- catch
        # both, since Drive doesn't always keep it zipped.
        if "geocene" in path_lower and (base.lower().endswith(".zip") or base.lower() == "missions.csv"):
            m_house = re.search(r"[Hh](\d+)\s*(PRE|POST)", path)
            if m_house:
                key = (m_house.group(1), m_house.group(2).lower())
                source = path if base.lower().endswith(".zip") else os.path.dirname(path)
                if source not in geocene_raw_sources[key]:
                    geocene_raw_sources[key].append(source)
            continue

        m = match_filename(base)
        if not m:
            continue
        house_num, group_label, rest, ext = m
        key = (house_num, group_label)
        rest_l = rest.lower()
        ext_l = ext.lower()

        if rest_l.startswith("airassure"):
            if not _is_genuine_raw_airassure(path):
                print(f"AirAssure: {path} doesn't look like genuine raw data (header doesn't match) -- skipped")
                continue
            if "combined" in rest_l:
                airassure_combined_by_house[house_num].append((key, path))
            else:
                airassure_parts_by_house[house_num].append((key, path))
        elif rest_l.startswith("geocene_left"):
            if ext_l == "json" or "geocene_left" not in groups[key]:
                groups[key]["geocene_left"] = path
        elif rest_l.startswith("geocene_right"):
            if ext_l == "json" or "geocene_right" not in groups[key]:
                groups[key]["geocene_right"] = path
        elif rest_l.startswith("atmocube"):
            groups[key]["atmocube"] = path
        elif rest_l.startswith("kestrel"):
            groups[key]["kestrel"] = path
        elif rest_l.startswith("aranet"):
            groups[key]["aranet"] = path
        elif rest_l.startswith("anemometer"):
            groups[key]["anemometer"] = path
        elif rest_l.startswith("hobo"):
            groups[key]["hobo"] = path
        elif rest_l.startswith("ogawa"):
            groups[key]["ogawa"] = path
        elif rest_l.startswith("house_charac") or rest_l.startswith("housecharac"):
            groups[key]["house_charac"] = path

    all_houses = set(airassure_parts_by_house) | set(airassure_combined_by_house)
    for house_num in all_houses:
        parts = airassure_parts_by_house.get(house_num, [])
        combined = airassure_combined_by_house.get(house_num, [])
        for key, path in parts:
            groups[key].setdefault("airassure", []).append(path)
        part_ranges = [_airassure_date_range(p) for _, p in parts]
        for key, path in combined:
            crange = _airassure_date_range(path)
            if any(_ranges_overlap(crange, pr) for pr in part_ranges):
                print(f"AirAssure: {path} overlaps with individual part files -- skipped as a duplicate")
                continue
            groups[key].setdefault("airassure", []).append(path)

    # if no clean geocene_left/right.json was found for a group, fall back to the raw export
    for key, sources in geocene_raw_sources.items():
        if "geocene_left" not in groups[key] and "geocene_right" not in groups[key]:
            groups[key]["geocene_raw_sources"] = sources

    return groups


def resolve_input_path(path, temp_dirs):
    if os.path.isfile(path) and path.lower().endswith(".zip"):
        temp_dir = tempfile.mkdtemp(prefix="hbrl_")
        print(f"Extracting {path} -> {temp_dir} (temporary, removed after processing)")
        with zipfile.ZipFile(path) as zf:
            zf.extractall(temp_dir)
        temp_dirs.append(temp_dir)
        return temp_dir
    return path


def cleanup_temp_dirs(temp_dirs):
    for d in temp_dirs:
        shutil.rmtree(d, ignore_errors=True)


# Reads the raw Geocene export directly, used when the clean flattened json isn't there


def _read_geocene_raw_folder(folder):
    """missions.csv + metrics/*.json.gz. Side (left/right) comes from
    missions.csv's notes field."""
    side = None
    missions_path = os.path.join(folder, "missions.csv")
    if os.path.exists(missions_path):
        with open(missions_path, newline="") as f:
            for row in csv.DictReader(f):
                notes = (row.get("notes") or "").lower()
                if "left" in notes:
                    side = "left"
                elif "right" in notes:
                    side = "right"
                break

    metrics_dir = os.path.join(folder, "metrics")
    rows = []
    if os.path.isdir(metrics_dir):
        for fname in os.listdir(metrics_dir):
            mpath = os.path.join(metrics_dir, fname)
            with gzip.open(mpath) as f:
                data = json.load(f)
            for entry in data:
                if entry.get("sensor_type_id") != 1:  # 1 = k-type thermocouple, Celsius
                    continue
                raw_ts = entry.get("timestamp", "")
                val = entry.get("value")
                if val is None:
                    continue
                try:
                    dt_utc = datetime.strptime(raw_ts[:19], "%Y-%m-%dT%H:%M:%S")
                except ValueError:
                    continue
                var = f"stove_temp_{side}" if side else "stove_temp_unknown"
                rows.append((dt_utc, var, val, "C", "ok"))
    return side, rows


def parse_geocene_raw_export(source_path):
    """source_path can be a .zip file OR an already-extracted folder --
    Drive doesn't always keep these zipped, so both need to work."""
    if os.path.isdir(source_path):
        return _read_geocene_raw_folder(source_path)
    temp_dir = tempfile.mkdtemp(prefix="geocene_raw_")
    try:
        with zipfile.ZipFile(source_path) as zf:
            zf.extractall(temp_dir)
        return _read_geocene_raw_folder(temp_dir)
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


# Anemometer parsing (hood_airflow readings) + gap detection, shared helpers


def parse_anemometer_rows(path):
    rows = []
    with open(path, encoding="utf-8", errors="replace") as f:
        for line in f:
            raw = line.strip()
            if not raw:
                continue
            sep = "\t" if "\t" in raw else ","  # some exports use tabs instead of commas
            parts = [p.strip() for p in raw.split(sep)]
            if len(parts) < 4:
                continue
            date_time_field = parts[-1]
            if "," in date_time_field:
                date_s, time_s = [p.strip() for p in date_time_field.split(",", 1)]
            elif len(parts) >= 5:
                date_s, time_s = parts[-2], parts[-1]
            else:
                continue
            if not re.match(r"^\d{2}-\d{2}-\d{4}$", date_s):
                continue
            try:
                dt_local = datetime.strptime(f"{date_s} {time_s}", "%d-%m-%Y %H:%M:%S")
                value = float(parts[1])
            except ValueError:
                continue
            # No timezone in the file itself -- this is local Oregon time (PDT), so it gets converted to UTC
            dt = dt_local.replace(tzinfo=PDT).astimezone(timezone.utc).replace(tzinfo=None)
            rows.append((dt, value))
    return rows


def detect_gaps(label, house_num, week_num, timestamps, min_gap_minutes=15):
    ts_sorted = sorted(set(timestamps))
    if len(ts_sorted) < 2:
        return
    diffs_sec = sorted((b - a).total_seconds() for a, b in zip(ts_sorted, ts_sorted[1:]))
    median = diffs_sec[len(diffs_sec) // 2]
    threshold = max(median * 3, min_gap_minutes * 60)
    for a, b in zip(ts_sorted, ts_sorted[1:]):
        gap = (b - a).total_seconds()
        if gap > threshold:
            a_pdt, b_pdt = a - timedelta(hours=7), b - timedelta(hours=7)
            date_str = a_pdt.strftime("%m/%d") if a_pdt.date() == b_pdt.date() else f"{a_pdt.strftime('%m/%d')}-{b_pdt.strftime('%m/%d')}"
            print(
                f"House {house_num} Week {week_num} {label}: no data on {date_str} "
                f"between {fmt_time(a_pdt)} and {fmt_time(b_pdt)} PDT"
            )


def parse_flexible_datetime(raw):
    raw = raw.strip()
    for fmt in ["%Y-%m-%d %I:%M:%S %p", "%m/%d/%y %I:%M:%S %p", "%m/%d/%Y %I:%M:%S %p", "%Y-%m-%d %H:%M:%S"]:
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    return None


# Per-instrument parsers. Each returns a list of (dt_utc_naive, variable, value, unit, qc_flag)


ATMOCUBE_MAP = {
    "voc": ("voc", "index"), "pm1": ("pm1", "ug/m3"), "pm25": ("pm25", "ug/m3"),
    "pm4": ("pm4", "ug/m3"), "pm10": ("pm10", "ug/m3"), "co2": ("co2", "ppm"),
    "t": ("temperature", "C"), "h": ("humidity", "%"), "abs_h": ("absolute_humidity", "g/m3"),
    "p": ("pressure", "hPa"), "noise": ("noise", "dB"), "light": ("light", "lux"),
    "ch2o": ("ch2o", "ppb"), "voc_index": ("voc_index", "index"), "nox_index": ("nox_index", "index"),
}


def parse_atmocube(path):
    """Uses the 'ts' unix epoch column."""
    rows = []
    with open(path, newline="") as f:
        for r in csv.DictReader(f):
            ts = r.get("ts")
            if not ts:
                continue
            try:
                epoch = int(ts)
            except ValueError:
                continue
            dt_utc = datetime.fromtimestamp(epoch, tz=timezone.utc).replace(tzinfo=None)
            for col, (var, unit) in ATMOCUBE_MAP.items():
                val = r.get(col)
                if val in (None, ""):
                    continue
                rows.append((dt_utc, var, val, unit, "ok"))
    return rows


def parse_geocene(path, side):
    """If the raw timestamp ends in 'Z', it's already UTC -- no conversion.
    If not, treat it as local Oregon time (PDT) and convert."""
    rows = []
    var = f"stove_temp_{side}"
    with open(path) as f:
        data = json.load(f)
    for entry in data:
        raw_ts = entry.get("timestamp", "")
        val = entry.get("value")
        if val is None:
            continue
        is_utc = raw_ts.strip().upper().endswith("Z")
        try:
            if is_utc:
                dt_utc = datetime.strptime(raw_ts[:19], "%Y-%m-%dT%H:%M:%S")
            else:
                dt_local = datetime.strptime(raw_ts[:19], "%Y-%m-%dT%H:%M:%S")
                dt_utc = dt_local.replace(tzinfo=PDT).astimezone(timezone.utc).replace(tzinfo=None)
        except ValueError:
            continue
        rows.append((dt_utc, var, val, "C", "ok"))
    return rows


def parse_hobo(path):
    """Header labels the offset (e.g. 'GMT-07:00') -- converted directly to UTC."""
    rows = []
    hobo_tz = None
    header_row = None
    with open(path, newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            joined = ",".join(row)
            if hobo_tz is None:
                m = re.search(r"GMT([+-]\d{2}):?(\d{2})", joined)
                if m:
                    sign = -1 if m.group(1).startswith("-") else 1
                    hours = int(m.group(1).lstrip("+-"))
                    hobo_tz = timezone(timedelta(hours=sign * hours))
            if header_row is None and any("Date" in c or "Time" in c for c in row):
                header_row = row
                continue
            if header_row is None:
                continue
            try:
                ts_col = next(i for i, c in enumerate(header_row) if "Date" in c or "Time" in c)
                pow_col = next(i for i, c in enumerate(header_row) if "Power" in c)
                raw_ts = row[ts_col]
                raw_pow = row[pow_col]
                dt_local = datetime.strptime(raw_ts.strip(), "%m/%d/%y %I:%M:%S %p")
            except (StopIteration, ValueError, IndexError):
                continue
            tz_used = hobo_tz or timezone(timedelta(hours=-7))
            dt_utc = dt_local.replace(tzinfo=tz_used).astimezone(timezone.utc).replace(tzinfo=None)
            if raw_pow not in (None, ""):
                rows.append((dt_utc, "power", raw_pow, "W", "ok"))
    return rows


def parse_kestrel(path):
    """No timezone in the file itself -- this is local Oregon time (PDT), so it gets converted to UTC."""
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        lines = f.readlines()

    header_idx = None
    for i, line in enumerate(lines):
        if re.search(r"temperature", line, re.IGNORECASE):
            header_idx = i
            break
    if header_idx is None:
        return rows

    reader = csv.DictReader(lines[header_idx:])
    header = reader.fieldnames or []

    ts_col = next((c for c in header if c and "timestamp" in c.lower()), None)
    if ts_col is None:
        ts_col = next((c for c in header if c and "date" in c.lower() and "time" in c.lower()), None)

    temp_c_col = next((c for c in header if c and "temperature" in c.lower() and ("(c)" in c.lower() or "\u00b0c" in c.lower())), None)
    temp_f_col = next((c for c in header if c and "temperature" in c.lower() and temp_c_col != c), None)
    hum_col = next((c for c in header if c and "humidity" in c.lower()), None)

    if ts_col is None:
        return rows

    for r in reader:
        raw_ts = r.get(ts_col)
        if not raw_ts:
            continue
        dt_local = parse_flexible_datetime(raw_ts)
        if dt_local is None:
            continue
        dt = dt_local.replace(tzinfo=PDT).astimezone(timezone.utc).replace(tzinfo=None)

        temp_c = r.get(temp_c_col) if temp_c_col else None
        if temp_c in (None, "") and temp_f_col and r.get(temp_f_col) not in (None, ""):
            try:
                temp_c = (float(r[temp_f_col]) - 32) * 5 / 9
            except ValueError:
                temp_c = None
        if temp_c not in (None, ""):
            rows.append((dt, "temperature", temp_c, "C", "ok"))

        hum = r.get(hum_col) if hum_col else None
        if hum not in (None, ""):
            rows.append((dt, "humidity", hum, "%", "ok"))
    return rows


def parse_aranet(path):
    """No timezone in the file itself -- this is local Oregon time (PDT), so it gets converted to UTC."""
    rows = []
    with open(path, newline="") as f:
        for r in csv.DictReader(f):
            raw_ts = list(r.values())[0]
            try:
                dt_local = datetime.strptime(raw_ts.strip(), "%d/%m/%Y %I:%M:%S %p")
            except (ValueError, AttributeError):
                continue
            dt = dt_local.replace(tzinfo=PDT).astimezone(timezone.utc).replace(tzinfo=None)
            co2 = r.get("Carbon dioxide(ppm)")
            if co2 not in (None, ""):
                rows.append((dt, "co2", co2, "ppm", "ok"))
            temp_f = r.get("Temperature(\u00b0F)") or r.get("Temperature(F)")
            temp_c = r.get("Temperature(\u00b0C)") or r.get("Temperature(C)")
            if temp_c not in (None, ""):
                rows.append((dt, "temperature", temp_c, "C", "ok"))
            elif temp_f not in (None, ""):
                try:
                    rows.append((dt, "temperature", (float(temp_f) - 32) * 5 / 9, "C", "ok"))
                except ValueError:
                    pass
            hum = r.get("Relative humidity(%)")
            if hum not in (None, ""):
                rows.append((dt, "humidity", hum, "%", "ok"))
            pres = r.get("Atmospheric pressure(hPa)")
            if pres not in (None, ""):
                rows.append((dt, "pressure", pres, "hPa", "ok"))
    return rows


GENERAL_SENSOR_CODES = {
    1: "not_factory_calibrated", 2: "hardware_fault", 4: "communication_error",
    8: "data_corrupt", 16: "data_not_available", 32: "over_detectable_limit",
    64: "under_detectable_limit",
}
PM_SENSOR_CODES = {**GENERAL_SENSOR_CODES, 256: "fan_rpm_error", 512: "laser_error",
                    1024: "fan_blocked", 2048: "cleaning_cycle_completed"}
VOC_SENSOR_CODES = {**GENERAL_SENSOR_CODES, 256: "sensor_unsupported"}
SYSTEM_STATUS_BITS = {
    1: "device_rebooted", 2: "data_not_written_to_sd", 4: "cloud_disconnected",
    8: "battery_low", 16: "time_not_synced", 32: "time_invalid", 64: "eeprom_failure",
}


def _normalize_col(name):
    """Strips anything in parens (units), then removes all non-alphanumeric
    characters and lowercases -- makes column matching robust across the
    different AirAssure header conventions seen in real deliveries, e.g.
    'PM_25(ug/m3)' and 'PM2.5' both normalize to 'pm25'."""
    name = re.sub(r"\([^)]*\)", "", name)
    return re.sub(r"[^a-z0-9]", "", name.lower())


# each entry maps a column name from the file -> what we call it, which error codes apply, and where to find its status column
AIRASSURE_VAR_MAP = {
    "temperature": ("temperature", GENERAL_SENSOR_CODES, ["temperaturestatus", "temperaturerelativehumiditystatus", "relativehumiditystatus"]),
    "relativehumidity": ("humidity", GENERAL_SENSOR_CODES, ["relativehumiditystatus", "temperaturerelativehumiditystatus"]),
    "barometricpressure": ("pressure", GENERAL_SENSOR_CODES, ["barometricpressurestatus"]),
    "pm1": ("pm1", PM_SENSOR_CODES, ["pmstatus"]),
    "pm25": ("pm25", PM_SENSOR_CODES, ["pmstatus"]),
    "pm4": ("pm4", PM_SENSOR_CODES, ["pmstatus"]),
    "pm10": ("pm10", PM_SENSOR_CODES, ["pmstatus"]),
    "co2": ("co2", GENERAL_SENSOR_CODES, ["co2status"]),
    "co": ("co", GENERAL_SENSOR_CODES, ["costatus"]),
    "no2": ("no2", GENERAL_SENSOR_CODES, ["no2status"]),
    "o3": ("o3", GENERAL_SENSOR_CODES, ["o3status"]),
    "so2": ("so2", GENERAL_SENSOR_CODES, ["so2status"]),
    "etoh": ("etoh", VOC_SENSOR_CODES, ["etohstatus", "vocstatus"]),
    "tvoc": ("tvoc", VOC_SENSOR_CODES, ["vocstatus", "tvocstatus"]),
}


def decode_bits(value, code_map):
    if not value:
        return []
    try:
        value = int(value)
    except (TypeError, ValueError):
        return []
    meanings = []
    for code, meaning in sorted(code_map.items(), reverse=True):
        if value & code:
            meanings.append(meaning)
    remaining = value
    for code in code_map:
        remaining &= ~code
    if remaining:
        meanings.append(f"unrecognized_status_bit_{remaining}")
    return meanings


def _parse_airassure_timestamp(raw_ts):
    """Different deliveries use different timestamp formats for this column."""
    for fmt in ("%m/%d/%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y %H:%M:%S"):
        try:
            return datetime.strptime(raw_ts, fmt)
        except ValueError:
            continue
    return None


def parse_airassure(paths, house_num, week_num):
    """File's units row literally says UTC, so no conversion needed."""
    header = None
    units_row = None
    data_lines = []
    for p in sorted(paths):
        with open(p, newline="") as f:
            for raw in f:
                line = raw.rstrip("\n")
                if not line.strip() or line.startswith("#"):
                    continue
                if line.startswith("Timestamp,System Status"):
                    if header is None:
                        header = next(csv.reader([line]))
                    continue
                if line.startswith("UTC,"):
                    if units_row is None and header is not None:
                        units_row = dict(zip(header, next(csv.reader([line]))))
                    continue
                data_lines.append(line)

    if header is None:
        print(f"House {house_num} Week {week_num} AirAssure: no recognizable header found, skipping")
        return []

    normalized_header = {_normalize_col(h): h for h in header}
    resolved = {}  # actual_col_name -> (our_var_name, code_map, actual_status_col_name_or_None, unit)
    unresolved = []
    for norm_key, (var_name, code_map, status_candidates) in AIRASSURE_VAR_MAP.items():
        actual_col = normalized_header.get(norm_key)
        if actual_col is None:
            unresolved.append(var_name)
            continue
        actual_status_col = next((normalized_header[c] for c in status_candidates if c in normalized_header), None)
        unit = (units_row or {}).get(actual_col, "")
        resolved[actual_col] = (var_name, code_map, actual_status_col, unit)
    if unresolved:
        print(f"House {house_num} Week {week_num} AirAssure: couldn't find a column for: {', '.join(unresolved)} (header may use an unrecognized naming convention)")

    rows = []
    seen_ts = set()
    all_ts = []
    ts_errors = defaultdict(list)

    for parts in csv.reader(data_lines):
        rec = dict(zip(header, parts))
        raw_ts = rec.get("Timestamp")
        if not raw_ts:
            continue
        dt_utc = _parse_airassure_timestamp(raw_ts)
        if dt_utc is None:
            continue
        if dt_utc in seen_ts:
            continue
        seen_ts.add(dt_utc)
        all_ts.append(dt_utc)

        try:
            sys_status = int(rec.get("System Status") or 0)
        except ValueError:
            sys_status = 0
        sys_flags = decode_bits(sys_status, SYSTEM_STATUS_BITS)
        sys_flags_reportable = [f for f in sys_flags if f != "cloud_disconnected"]
        if sys_flags_reportable:
            ts_errors[dt_utc].append("system: " + "+".join(sys_flags_reportable))

        for col, (var, code_map, status_col, unit) in resolved.items():
            val = rec.get(col)
            if val in (None, ""):
                continue
            try:
                status_val = int(rec.get(status_col) or 0) if status_col else 0
            except ValueError:
                status_val = 0
            flags = decode_bits(status_val, code_map)
            if status_val != 0 and f"{var}: " + "+".join(flags) not in ts_errors[dt_utc]:
                ts_errors[dt_utc].append(f"{var}: " + "+".join(flags))
            qc_parts = flags + sys_flags_reportable
            qc = "+".join(qc_parts) if qc_parts else "ok"
            rows.append((dt_utc, var, val, unit, qc))

    for dt_utc in sorted(ts_errors):
        reasons = "; ".join(ts_errors[dt_utc])
        dt_pdt_display = dt_utc - timedelta(hours=7)
        print(
            f"House {house_num} Week {week_num} AirAssure at {fmt_time(dt_pdt_display)} on "
            f"{dt_pdt_display.strftime('%Y-%m-%d')} PDT had an error: {reasons}"
        )

    detect_gaps("AirAssure", house_num, week_num, all_ts)
    return rows


# House characteristics table (static per-house metadata, not a timeseries)


HOUSE_CHARAC_COLUMN_MAP = {
    "House": "house_id", "Visit 1": "visit_1_date", "Visit 2": "visit_2_date", "Visit 3": "visit_3_date",
    "Gas Stove Make": "gas_stove_make", "Gas Stove Model": "gas_stove_model", "Oven Type": "oven_type",
    "Gas Stove Approx. Age (years)": "gas_stove_age_years", "Gas Stove Number of Burners": "gas_stove_burners",
    "Range Hood Type": "range_hood_type", "Range Hood Make": "range_hood_make", "Range Hood Model": "range_hood_model",
    "Range Hood Age (years)": "range_hood_age_years",
    "Range Hood Velocity - Left - Off (m/s)": "hood_velocity_left_off",
    "Range Hood Velocity - Left - High (m/s)": "hood_velocity_left_high",
    "Range Hood Velocity - Left - Low (m/s)": "hood_velocity_left_low",
    "Range Hood Velocity - Right - Off (m/s)": "hood_velocity_right_off",
    "Range Hood Velocity - Right - High (m/s)": "hood_velocity_right_high",
    "Range Hood Velocity - Right - Low (m/s)": "hood_velocity_right_low",
    "Type of House": "house_type", "Kitchen Area (sq-ft)": "kitchen_area_sqft",
    "Kitchen Height (ft)": "kitchen_height_ft", "House Area (sq-ft)": "house_area_sqft",
    "Number of Floors": "num_floors", "Number of Bedrooms": "num_bedrooms",
    "Members of Household": "num_household_members", "Other Sources of NO2": "other_no2_sources",
}


def build_house_characteristics(source_file, output_folder, house_filter=None):
    if openpyxl is None:
        print("openpyxl is required for this -- install it with: pip install openpyxl")
        sys.exit(1)
    wb = openpyxl.load_workbook(source_file, data_only=True)
    ws = wb["Sheet1"]
    header_row = [cell.value for cell in ws[2]]

    col_indexes = {}
    for idx, raw_header in enumerate(header_row):
        if raw_header in HOUSE_CHARAC_COLUMN_MAP:
            col_indexes[HOUSE_CHARAC_COLUMN_MAP[raw_header]] = idx
        elif raw_header is not None:
            print(f"Note: unmapped House_charac column found: {raw_header!r} -- skipped")

    rows_out = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue
        record = {}
        for clean_name, idx in col_indexes.items():
            val = row[idx] if idx < len(row) else None
            if clean_name.startswith("visit_") and val is not None:
                val = val.strftime("%Y-%m-%d %H:%M:%S")
            record[clean_name] = val if val is not None else ""
        if house_filter and record.get("house_id") not in house_filter:
            continue
        rows_out.append(record)

    out_path = os.path.join(output_folder, "house_characteristics.csv")
    fieldnames = list(HOUSE_CHARAC_COLUMN_MAP.values())
    with open(out_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows_out)

    print(f"Wrote {len(rows_out)} house(s) to {out_path}")
    return out_path


# Ogawa NO2 table (weekly integrated badge samples, not a timeseries either)


OGAWA_RAW_COL_TO_LOCATION = {
    "Ogawa_1": "kitchen_near_stove", "Ogawa_2": "kitchen_far_stove", "Ogawa_3": "living_room",
    "Ogawa_4": "bedroom", "Ogawa_5_Outside": "outdoors", "Ogawa_6_FB": "field_blank", "Ogawa_7_LB": "lab_blank",
}
OGAWA_PLACEMENT_COL_TO_LOCATION = {
    "F": "kitchen_near_stove", "H": "kitchen_far_stove", "J": "living_room", "L": "outdoors", "N": "bedroom",
}


def _load_ogawa_placement_descriptions(placement_file):
    wb = openpyxl.load_workbook(placement_file, data_only=True)
    ws = wb["Placement"]
    descriptions = {}
    for row in ws.iter_rows(min_row=2, values_only=False):
        house_raw = row[0].value
        if not house_raw:
            continue
        house_id = str(house_raw).strip()
        if house_id in descriptions:
            continue
        row_map = {cell.coordinate[0]: cell.value for cell in row}
        descriptions[house_id] = {loc_key: row_map.get(col, "") for col, loc_key in OGAWA_PLACEMENT_COL_TO_LOCATION.items()}
    return descriptions


def build_ogawa_table(values_file, placement_file, output_folder, house_filter=None):
    if openpyxl is None:
        print("openpyxl is required for this -- install it with: pip install openpyxl")
        sys.exit(1)
    wb = openpyxl.load_workbook(values_file, data_only=True)
    ws = wb["Ogawa_raw"]
    placement = _load_ogawa_placement_descriptions(placement_file)

    header = [cell.value for cell in ws[1]]
    rows_out = []
    skipped = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(header, row))
        house_raw = record.get("HouseID")
        phase_raw = record.get("Pre_Post")
        if house_raw is None or phase_raw is None:
            continue
        house_id = f"H{int(house_raw)}"
        phase = str(phase_raw).strip().lower()
        notes = record.get("Notes") or ""
        if house_filter and house_id not in house_filter:
            continue

        loc_descriptions = placement.get(house_id, {})
        had_any_value = False
        for col, loc_key in OGAWA_RAW_COL_TO_LOCATION.items():
            val = record.get(col)
            if val is None or val == "":
                continue
            had_any_value = True
            rows_out.append({
                "house_id": house_id, "phase": phase, "location": loc_key,
                "location_description": loc_descriptions.get(loc_key, ""),
                "value_ppb": val, "notes": notes,
            })
        if not had_any_value:
            skipped.append((house_id, phase))

    out_path = os.path.join(output_folder, "ogawa.csv")
    fieldnames = ["house_id", "phase", "location", "location_description", "value_ppb", "notes"]
    with open(out_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows_out)

    print(f"Wrote {len(rows_out)} rows to {out_path}")
    print(f"Houses found: {sorted(set(r['house_id'] for r in rows_out))}")
    if skipped:
        print(f"Skipped {len(skipped)} house/phase combo(s) with no results yet: {skipped}")
    return out_path


# Sensor IDs (static per-house metadata, added as a column on the sensors table)


SENSOR_ID_COLUMN_MAP = {
    "Atmocube": "Atmocube (final ID)", "AirAssure": "Air Assure (Final ID)", "Aranet": "Aranet",
    "Anemometer": "Anemometer DigiSense", "Kestrel": "Kestrel", "Hobo": "Ind Stove - HOBO Plug Monitor",
}
GEOCENE_SENSOR_ID_COLUMN_MAP = {"stove_temp_left": "Geocene Left", "stove_temp_right": "Geocene Right"}


def load_sensor_ids(path):
    """Geocene is keyed by variable (stove_temp_left/right), not
    instrument, since both burners need different sensor IDs."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    all_cols = {**SENSOR_ID_COLUMN_MAP, **GEOCENE_SENSOR_ID_COLUMN_MAP}
    col_idx = {name: header.index(name) for name in all_cols.values() if name in header}

    sensor_ids = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        house_raw = row[0]
        if not house_raw or not str(house_raw).strip().upper().startswith("H") or "&" in str(house_raw):
            continue
        house_id = str(house_raw).strip()
        sensor_ids[house_id] = {key: row[col_idx[col]] for key, col in all_cols.items() if col in col_idx}
    return sensor_ids


# Visit dates / phase windows


def load_visit_dates(house_charac_file):
    """Returns house_id -> [visit1_dt, visit2_dt, visit3_dt or None], all in true UTC."""
    wb = openpyxl.load_workbook(house_charac_file, data_only=True)
    ws = wb["Sheet1"]
    header_row = [cell.value for cell in ws[2]]
    idx = {name: i for i, name in enumerate(header_row) if name in ("House", "Visit 1", "Visit 2", "Visit 3")}

    def to_utc(v):
        return v.replace(tzinfo=PDT).astimezone(timezone.utc).replace(tzinfo=None) if v else None

    visits = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        house_id = row[idx["House"]] if "House" in idx else None
        if not house_id:
            continue
        v1 = to_utc(row[idx["Visit 1"]] if "Visit 1" in idx else None)
        v2 = to_utc(row[idx["Visit 2"]] if "Visit 2" in idx else None)
        v3 = to_utc(row[idx["Visit 3"]] if "Visit 3" in idx else None)
        visits[str(house_id).strip()] = (v1, v2, v3)
    return visits


def compute_phase_windows(v1, v2, v3, pre_buffer=timedelta(minutes=30), post_buffer=timedelta(hours=1)):
    pre_window = (v1, v2 - pre_buffer) if v1 and v2 else None
    post_window = (v2 + post_buffer, v3) if v2 and v3 else None
    return pre_window, post_window


def tag_phase(dt, pre_window, post_window):
    if pre_window and pre_window[0] <= dt <= pre_window[1]:
        return "pre"
    if post_window and post_window[0] <= dt <= post_window[1]:
        return "post"
    return None


def process_house(house_num, week_groups, house_charac_file, output_folder, sensor_ids=None):
    house_id = f"H{house_num}"
    print(f"\n=== Processing House {house_num} (groups found: {sorted(week_groups)}) ===")

    visits = load_visit_dates(house_charac_file)
    if house_id not in visits:
        print(f"House {house_num}: not found in {house_charac_file} -- skipping, no visit dates to work from")
        return None
    v1, v2, v3 = visits[house_id]
    if v1 is None or v2 is None or v3 is None:
        print(f"House {house_num}: missing one or more visit dates (v1={v1}, v2={v2}, v3={v3}) -- can't compute pre/post, skipping")
        return None

    pre_window, post_window = compute_phase_windows(v1, v2, v3)
    print("  PRE/POST WINDOWS (from House_charac Visit 1/2/3, converted to true UTC):")
    print(f"    pre:  {pre_window[0]} UTC  ->  {pre_window[1]} UTC")
    print(f"    post: {post_window[0]} UTC  ->  {post_window[1]} UTC")
    print(f"    (visit + buffer excluded: {pre_window[1]} -> {post_window[0]} UTC)")
    print("-" * 42)

    out_rows = []
    counts = defaultdict(lambda: defaultdict(int))

    def add(instrument, parsed):
        for dt, var, val, unit, qc in parsed:
            phase = tag_phase(dt, pre_window, post_window)
            if phase is None:
                counts[instrument]["dropped_buffer_or_outside"] += 1
                continue
            counts[instrument][phase] += 1
            out_rows.append((dt, phase, instrument, var, val, unit, qc))

    for week_num, files in sorted(week_groups.items()):
        if "atmocube" in files:
            add("Atmocube", parse_atmocube(files["atmocube"]))
        if "kestrel" in files:
            add("Kestrel", parse_kestrel(files["kestrel"]))
        if "aranet" in files:
            add("Aranet", parse_aranet(files["aranet"]))
        if "geocene_left" in files:
            add("Geocene", parse_geocene(files["geocene_left"], "left"))
        if "geocene_right" in files:
            add("Geocene", parse_geocene(files["geocene_right"], "right"))
        if "geocene_raw_sources" in files:
            for source in files["geocene_raw_sources"]:
                side, rows = parse_geocene_raw_export(source)
                print(f"House {house_num} group {week_num}: recovered {len(rows)} Geocene rows ({side or 'unknown side'}) from raw export {source}")
                add("Geocene", rows)
        if "hobo" in files:
            add("Hobo", parse_hobo(files["hobo"]))
        if "anemometer" in files:
            anem_rows = parse_anemometer_rows(files["anemometer"])
            add("Anemometer", [(dt, "hood_airflow", val, "m/s", "ok") for dt, val in anem_rows])
        if "airassure" in files:
            airassure_rows = parse_airassure(files["airassure"], house_num, week_num)
            add("AirAssure", airassure_rows)
        if "ogawa" in files:
            print(f"House {house_num} group {week_num}: Ogawa file present here -- use the 'ogawa' subcommand instead, not included in this table")
        if "house_charac" in files:
            print(f"House {house_num} group {week_num}: House_charac file present here -- already used above for visit dates, not included as timeseries rows")

    print("Rows per instrument (pre / post / dropped as visit-buffer or outside deployment):")
    for instrument, c in sorted(counts.items()):
        print(f"  {instrument}: pre={c['pre']}  post={c['post']}  dropped={c['dropped_buffer_or_outside']}")

    seen = set()
    deduped_rows = []
    dupes_removed = defaultdict(int)
    for dt, phase, instrument, var, val, unit, qc in out_rows:
        key = (dt, instrument, var)
        if key in seen:
            dupes_removed[instrument] += 1
            continue
        seen.add(key)
        deduped_rows.append((dt, phase, instrument, var, val, unit, qc))
    if dupes_removed:
        print("Duplicate readings removed (same timestamp+instrument+variable found in more than one source file):")
        for instrument, n in sorted(dupes_removed.items()):
            print(f"  {instrument}: {n} duplicate(s) removed")
    out_rows = deduped_rows
    out_rows.sort(key=lambda r: (r[0], r[2], r[3]))

    out_path = os.path.join(output_folder, f"{house_id}.csv")
    fields = ["house_id", "phase", "date", "timestamp_utc", "instrument", "variable", "value", "unit", "qc_flag", "sensor_id"]
    house_sensor_ids = (sensor_ids or {}).get(house_id, {})
    with open(out_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(fields)
        for dt, phase, instrument, var, val, unit, qc in out_rows:
            if instrument == "Geocene":
                sensor_id = house_sensor_ids.get(var, "")
            else:
                sensor_id = house_sensor_ids.get(instrument, "")
            writer.writerow([house_id, phase, dt.strftime("%Y-%m-%d"), dt.strftime("%Y-%m-%d %H:%M:%S"), instrument, var, val, unit, qc, sensor_id])

    print(f"Wrote {len(out_rows)} rows to {out_path}")
    return out_path


# verify-timezones: audits the ACTUAL production parser functions above


TZ_SOURCE_NOTES = {
    "Atmocube": ("Unix epoch, already UTC", "datetime.fromtimestamp(epoch, tz=UTC)"),
    "Geocene": ("Ends in 'Z' = UTC; otherwise treated as local PDT", "Parsed directly if 'Z', else convert (+7h)"),
    "AirAssure": ("File says UTC in the units row", "Parsed directly, no conversion"),
    "Hobo": ("File header gives the offset (e.g. GMT-07:00)", "Convert using that offset"),
    "Kestrel": ("Local Oregon time, PDT (UTC-7)", "Convert to UTC (+7h)"),
    "Aranet": ("Local Oregon time, PDT (UTC-7)", "Convert to UTC (+7h)"),
    "Anemometer": ("Local Oregon time, PDT (UTC-7)", "Convert to UTC (+7h)"),
}


def _peek_raw_timestamp(kind, path, n):
    raw_values = []
    try:
        if kind == "atmocube":
            with open(path, newline="") as f:
                for r in csv.DictReader(f):
                    if r.get("ts"):
                        raw_values.append(r["ts"])
                    if len(raw_values) >= n:
                        break
        elif kind == "geocene":
            with open(path) as f:
                data = json.load(f)
            raw_values = [e.get("timestamp", "") for e in data[:n]]
        elif kind == "airassure":
            header = None
            for p in sorted(path):
                with open(p, newline="") as f:
                    for line in f:
                        line = line.rstrip("\n")
                        if not line.strip() or line.startswith("#"):
                            continue
                        if line.startswith("Timestamp,System Status"):
                            header = next(csv.reader([line]))
                            continue
                        if line.startswith("UTC,") or header is None:
                            continue
                        rec = dict(zip(header, next(csv.reader([line]))))
                        if rec.get("Timestamp"):
                            raw_values.append(rec["Timestamp"])
                        if len(raw_values) >= n:
                            return raw_values
        elif kind == "kestrel":
            with open(path, newline="", encoding="utf-8-sig") as f:
                lines = f.readlines()
            header_idx = next((i for i, line in enumerate(lines) if re.search(r"temperature", line, re.IGNORECASE)), None)
            if header_idx is not None:
                reader = csv.DictReader(lines[header_idx:])
                header = reader.fieldnames or []
                ts_col = next((c for c in header if c and "timestamp" in c.lower()), None)
                if ts_col is None:
                    ts_col = next((c for c in header if c and "date" in c.lower() and "time" in c.lower()), None)
                if ts_col:
                    for r in reader:
                        if r.get(ts_col):
                            raw_values.append(r[ts_col])
                        if len(raw_values) >= n:
                            break
        elif kind == "aranet":
            with open(path, newline="") as f:
                for r in csv.DictReader(f):
                    raw_values.append(list(r.values())[0])
                    if len(raw_values) >= n:
                        break
        elif kind == "anemometer":
            with open(path, encoding="utf-8", errors="replace") as f:
                for line in f:
                    raw = line.strip()
                    if not raw:
                        continue
                    sep = "\t" if "\t" in raw else ","
                    parts = [p.strip() for p in raw.split(sep)]
                    if len(parts) < 4:
                        continue
                    date_time_field = parts[-1]
                    if "," in date_time_field:
                        date_s, time_s = [p.strip() for p in date_time_field.split(",", 1)]
                    elif len(parts) >= 5:
                        date_s, time_s = parts[-2], parts[-1]
                    else:
                        continue
                    if not re.match(r"^\d{2}-\d{2}-\d{4}$", date_s):
                        continue
                    raw_values.append(f"{date_s} {time_s}")
                    if len(raw_values) >= n:
                        break
        elif kind == "hobo":
            with open(path, newline="") as f:
                for line in f:
                    if re.match(r"^\d{1,2}/\d{1,2}/\d{2}", line):
                        raw_values.append(line.split(",")[0])
                    if len(raw_values) >= n:
                        break
    except (FileNotFoundError, KeyError, IndexError, StopIteration):
        pass
    return raw_values


def build_timezone_report(input_path, samples=3, output_path=None):
    temp_dirs = []
    report_rows = []
    try:
        resolved = resolve_input_path(input_path, temp_dirs)
        groups = discover_groups(resolved)
        if not groups:
            print(f"No h<N>_(pre|post|w<N>)_... files found in {input_path}")
            return

        for (house_num, week_num), files in sorted(groups.items()):
            label = f"H{house_num}{week_num}"

            def add(instrument, parsed_rows, raw_values):
                seen_utc = []
                for dt_utc, *_ in parsed_rows:
                    if dt_utc not in seen_utc:
                        seen_utc.append(dt_utc)
                    if len(seen_utc) >= samples:
                        break
                tz_note, method = TZ_SOURCE_NOTES[instrument.split(" ")[0]]
                for raw, converted in zip(raw_values, seen_utc):
                    report_rows.append({
                        "group": label, "instrument": instrument, "raw_timestamp": raw,
                        "converted_utc": converted.strftime("%Y-%m-%d %H:%M:%S"),
                        "source_timezone_assumption": tz_note, "conversion_method": method,
                    })

            if "atmocube" in files:
                add("Atmocube", parse_atmocube(files["atmocube"]), _peek_raw_timestamp("atmocube", files["atmocube"], samples))
            if "geocene_left" in files:
                add("Geocene (left)", parse_geocene(files["geocene_left"], "left"), _peek_raw_timestamp("geocene", files["geocene_left"], samples))
            if "geocene_right" in files:
                add("Geocene (right)", parse_geocene(files["geocene_right"], "right"), _peek_raw_timestamp("geocene", files["geocene_right"], samples))
            if "hobo" in files:
                add("Hobo", parse_hobo(files["hobo"]), _peek_raw_timestamp("hobo", files["hobo"], samples))
            if "kestrel" in files:
                add("Kestrel", parse_kestrel(files["kestrel"]), _peek_raw_timestamp("kestrel", files["kestrel"], samples))
            if "aranet" in files:
                add("Aranet", parse_aranet(files["aranet"]), _peek_raw_timestamp("aranet", files["aranet"], samples))
            if "anemometer" in files:
                anem_rows = [(dt, None, None, None, None) for dt, val in parse_anemometer_rows(files["anemometer"])]
                add("Anemometer", anem_rows, _peek_raw_timestamp("anemometer", files["anemometer"], samples))
            if "airassure" in files:
                add("AirAssure", parse_airassure(files["airassure"], house_num, week_num), _peek_raw_timestamp("airassure", files["airassure"], samples))

        if not report_rows:
            print("No timestamp data found to report.")
            return

        cols = ["group", "instrument", "raw_timestamp", "converted_utc", "source_timezone_assumption", "conversion_method"]
        widths = {c: max(len(c), max(len(str(r[c])) for r in report_rows)) for c in cols}
        widths["source_timezone_assumption"] = min(widths["source_timezone_assumption"], 55)
        widths["conversion_method"] = min(widths["conversion_method"], 40)

        def line(vals):
            return " | ".join(str(v)[:widths[c]].ljust(widths[c]) for c, v in zip(cols, vals))

        print(line(cols))
        print("-+-".join("-" * widths[c] for c in cols))
        for r in report_rows:
            print(line([r[c] for c in cols]))

        out_path = output_path or "timezone_report.csv"
        with open(out_path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=cols)
            writer.writeheader()
            writer.writerows(report_rows)
        print(f"\nSaved {out_path}")
    finally:
        cleanup_temp_dirs(temp_dirs)


# Driver


DEFAULT_OUTPUT_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")


def build_arg_parser():
    parser = argparse.ArgumentParser(
        prog="process_hbrl_data.py",
        description="Build the HBRL standardized tables: sensor data, house characteristics, Ogawa NO2, or a timezone audit.",
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    sensors_p = subparsers.add_parser(
        "sensors",
        help="Process a house's raw instrument folders into ONE combined table, tagged pre/post using House_charac visit dates",
    )
    sensors_p.add_argument("house_charac_file", help="Path to the House Characteristics .xlsx file (used for pre/post windows)")
    sensors_p.add_argument("input_folders", nargs="+", help="One or more folders (or .zip files) for the SAME house, e.g. the pre folder/zip and the post folder/zip")
    sensors_p.add_argument("--sensor-ids", default=None, help="Path to the Sensors_used_in_houses.xlsx file (adds a sensor_id column)")
    sensors_p.add_argument("--output", default=None, help='Where to write H<N>.csv (default: an "output" folder next to this script)')

    house_p = subparsers.add_parser("house-charac", help="Build the house characteristics table from the House_charac spreadsheet")
    house_p.add_argument("source_file", help="Path to the House Characteristics .xlsx file")
    house_p.add_argument("output_folder", nargs="?", default=None, help='Where to write house_characteristics.csv (default: an "output" folder next to this script)')
    house_p.add_argument("--house", nargs="*", default=None, help="Only include specific house(s), e.g. --house H1 H2")

    ogawa_p = subparsers.add_parser("ogawa", help="Build the Ogawa NO2 table from the values + placement spreadsheets")
    ogawa_p.add_argument("values_file", help="Path to the Ogawa NO2 values .xlsx file")
    ogawa_p.add_argument("placement_file", help="Path to the Ogawa placement .xlsx file")
    ogawa_p.add_argument("output_folder", nargs="?", default=None, help='Where to write ogawa.csv (default: an "output" folder next to this script)')
    ogawa_p.add_argument("--house", nargs="*", default=None, help="Only include specific house(s), e.g. --house H1")

    tz_p = subparsers.add_parser("verify-timezones", help="Show raw vs. converted-UTC timestamps per instrument, for manual review")
    tz_p.add_argument("input_path", help="A folder or .zip file to audit")
    tz_p.add_argument("--samples", type=int, default=3, help="How many sample rows per instrument (default 3)")
    tz_p.add_argument("--output", default=None, help="CSV output path (default: timezone_report.csv)")

    return parser


def main():
    args = build_arg_parser().parse_args()

    if args.command == "sensors":
        temp_dirs = []
        try:
            all_groups = defaultdict(dict)
            for folder in args.input_folders:
                resolved = resolve_input_path(folder, temp_dirs)
                groups = discover_groups(resolved)
                for key, files in groups.items():
                    all_groups[key].update(files)
            if not all_groups:
                print(f"No h<N>_(pre|post|w<N>)_... files found in: {', '.join(args.input_folders)}")
                sys.exit(1)

            by_house = defaultdict(dict)
            for (house_num, group_label), files in all_groups.items():
                by_house[house_num][group_label] = files

            output_folder = args.output or DEFAULT_OUTPUT_FOLDER
            os.makedirs(output_folder, exist_ok=True)
            print(f"Writing output to {output_folder}")

            sensor_ids = load_sensor_ids(args.sensor_ids) if args.sensor_ids else None
            for house_num in sorted(by_house):
                process_house(house_num, by_house[house_num], args.house_charac_file, output_folder, sensor_ids=sensor_ids)
        finally:
            cleanup_temp_dirs(temp_dirs)

    elif args.command == "house-charac":
        output_folder = args.output_folder or DEFAULT_OUTPUT_FOLDER
        os.makedirs(output_folder, exist_ok=True)
        print(f"Writing output to {output_folder}")
        build_house_characteristics(args.source_file, output_folder, house_filter=args.house)

    elif args.command == "ogawa":
        output_folder = args.output_folder or DEFAULT_OUTPUT_FOLDER
        os.makedirs(output_folder, exist_ok=True)
        print(f"Writing output to {output_folder}")
        build_ogawa_table(args.values_file, args.placement_file, output_folder, house_filter=args.house)

    elif args.command == "verify-timezones":
        build_timezone_report(args.input_path, samples=args.samples, output_path=args.output)


if __name__ == "__main__":
    main()